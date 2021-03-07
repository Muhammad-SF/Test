from odoo import api, fields, models
from openpyxl import Workbook
from openpyxl.styles import Style, Font, Alignment, Border, Side
import tempfile
import base64
import os
import time
from datetime import date, datetime
from openerp.tools import DEFAULT_SERVER_DATE_FORMAT
import logging
_logger = logging.getLogger(__name__)


class AccountingReport(models.TransientModel):
    _inherit = 'accounting.report'

    account_report_name = fields.Char(related='account_report_id.name',string='Report Name')

    @api.multi
    @api.onchange('date_from','date_to')
    def onchage_check_date(self):
        warning = {}
        if (self.date_from and self.date_to) and (self.date_from > self.date_to):
            self.date_to = False
            warning = {'message': "Value Error\nEnd Date Should Be Greater Than Start Date"}
        return {'warning': warning}

    @api.multi
    def action_export(self):
        context=dict(acc_id=self.id)
        return {
            'name': 'Generate Excel Wizard',
            'view_type': 'form',
            'view_mode': 'form',
            'target': 'new',
            'res_model': 'cash.flow.statement',
            'type': 'ir.actions.act_window',
            'context' : context
        }

AccountingReport()

class CashFlowStatement(models.TransientModel):
    _name = 'cash.flow.statement'
    # _inherit = 'report.account.report_financial'

    file = fields.Binary(string="File")
    filename = fields.Char(string='Filename')

    @api.multi
    def action_generate_xls(self):
        current_lang_date_format = '%d/%m/%Y'
        lang_obj = self.env['res.lang'].search([('code','=',self._context.get('lang'))])
        if lang_obj:
           current_lang_date_format = lang_obj.date_format
        excel_fd, excelreport_path = tempfile.mkstemp(suffix='.xlsx', prefix='excel.tmp.')
        wb= Workbook()
        ws= wb.active
        context = dict(self._context or {})
        if context.get('active_ids'):
            report_obj = self.env['accounting.report'].browse(context.get('active_ids'))
            report_ob = report_obj[0]
            entries_dict = {'all': 'All Entries', 'posted': 'All Posted Entries'}
            if report_ob.account_report_id.name == 'Statement Of Cash Flow':
                ws.append([self.env.user.company_id.name])
                ws.append(['Statement Of Cash Flow Report as of '+(datetime.today().strftime('%d/%m/%Y'))])
                ws.append(['Target Moves','','Date'])
                ws.append([entries_dict[report_ob.target_move],'',''+str(datetime.strptime(report_ob.date_from,DEFAULT_SERVER_DATE_FORMAT).strftime(current_lang_date_format)),''+str(datetime.strptime(report_ob.date_to,DEFAULT_SERVER_DATE_FORMAT).strftime(current_lang_date_format))])
                ws.append([''])
                ws.append([''])

                ws.append(['','','Jan','Feb','Mar','Apr','May','Jun','July','Aug','Sep','Oct','Nov','Dec','Balance'])
                ft = Font(size=12,bold=True)
                st = Style(font=ft)
                ws['A7'].style = st
                ws['A7'].alignment = Alignment(horizontal="center")
                ws.merge_cells('A7:B7')
                ws['C7'].style = st
                ws['C7'].alignment = Alignment(horizontal="center")
                ws['D7'].style = st
                ws['D7'].alignment = Alignment(horizontal="center")
                ws['E7'].style = st
                ws['E7'].alignment = Alignment(horizontal="center")
                ws['F7'].style = st
                ws['F7'].alignment = Alignment(horizontal="center")
                ws['G7'].style = st
                ws['G7'].alignment = Alignment(horizontal="center")
                ws['H7'].style = st
                ws['H7'].alignment = Alignment(horizontal="center")
                ws['I7'].style = st
                ws['I7'].alignment = Alignment(horizontal="center")
                ws['J7'].style = st
                ws['J7'].alignment = Alignment(horizontal="center")
                ws['K7'].style = st
                ws['K7'].alignment = Alignment(horizontal="center")
                ws['L7'].style = st
                ws['L7'].alignment = Alignment(horizontal="center")
                ws['M7'].style = st
                ws['M7'].alignment = Alignment(horizontal="center")
                ws['N7'].style = st
                ws['N7'].alignment = Alignment(horizontal="center")
                ws['O7'].style = st
                ws['O7'].alignment = Alignment(horizontal="center")
                thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=7, column=1).border = thin_border
                ws.cell(row=7, column=2).border = thin_border
                ws.cell(row=7, column=3).border = thin_border
                ws.cell(row=7, column=4).border = thin_border
                ws.cell(row=7, column=5).border = thin_border
                ws.cell(row=7, column=6).border = thin_border
                ws.cell(row=7, column=7).border = thin_border
                ws.cell(row=7, column=8).border = thin_border
                ws.cell(row=7, column=9).border = thin_border
                ws.cell(row=7, column=10).border = thin_border
                ws.cell(row=7, column=11).border = thin_border
                ws.cell(row=7, column=12).border = thin_border
                ws.cell(row=7, column=13).border = thin_border
                ws.cell(row=7, column=14).border = thin_border
                ws.cell(row=7, column=15).border = thin_border

                # Width style
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 15
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 15
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 15
                ws.column_dimensions['L'].width = 15
                ws.column_dimensions['M'].width = 15
                ws.column_dimensions['N'].width = 15
                ws.column_dimensions['O'].width = 15
                t_count = 7

                domain=[]
                if report_ob.target_move == 'posted':
                    domain.append(('move_id.state','=','posted'))
                elif report_ob.target_move == 'all':
                    domain.append(('move_id.state','in',['draft','posted']))

                #Balance For View type
                report_view_obj = self.env['account.financial.report'].search([('parent_id','=',report_ob.account_report_id.id),('type','=','sum'),('sign','=','1')])
                for view_report in report_view_obj:
                    view_jan_bal,view_feb_bal,view_mar_bal,view_apr_bal,view_may_bal,view_jun_bal,view_july_bal=0.00,0.00,0.00,0.00,0.00,0.00,0.00
                    view_aug_bal,view_sep_bal,view_oct_bal,view_nov_bal,view_dec_bal = 0.00,0.00,0.00,0.00,0.00
                    report_obj = self.env['account.financial.report'].search([('parent_id','=',view_report.id),('type','=','account_type'),('sign','=','-1')])
                    for vreport in report_obj:
                        vreport_ids = self.env['account.account'].search([('user_type_id','in', [x.id for x in vreport.account_type_ids])])
                        for view_acc in vreport_ids:
                            year_from = report_ob.date_from[:4]
                            year_to = report_ob.date_to[:4]

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-01-01'),('date','<=',year_to+'-01-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_jan_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            if report_ob.date_to == year_to+'-02-29':
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-29'),('account_id','=',view_acc.id)]+domain)
                            else:
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-28'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_feb_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-03-01'),('date','<=',year_to+'-03-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_mar_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-04-01'),('date','<=',year_to+'-04-30'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_apr_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-05-01'),('date','<=',year_to+'-05-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_may_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-06-01'),('date','<=',year_to+'-06-30'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_jun_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-07-01'),('date','<=',year_to+'-07-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_july_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-08-01'),('date','<=',year_to+'-08-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_aug_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-09-01'),('date','<=',year_to+'-09-30'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_sep_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-10-01'),('date','<=',year_to+'-10-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_oct_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-11-01'),('date','<=',year_to+'-11-30'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_nov_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-12-01'),('date','<=',year_to+'-12-31'),('account_id','=',view_acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            view_dec_bal += debit_bal - credit_bal

                    view_tot_bal = view_jan_bal+ view_feb_bal+view_mar_bal+view_apr_bal+view_may_bal+view_jun_bal+view_july_bal+view_aug_bal+view_sep_bal+view_oct_bal+view_nov_bal+view_dec_bal

                    ws.append([''+view_report.name,'',view_jan_bal, view_feb_bal, view_mar_bal, view_apr_bal, view_may_bal, view_jun_bal, view_july_bal, view_aug_bal, view_sep_bal, view_oct_bal, view_nov_bal, view_dec_bal, view_tot_bal])
                    ft = Font(size=13, bold=True,color='A901DB')
                    st = Style(font=ft)
                    ws.merge_cells("A"+str(t_count+1)+":B"+str(t_count+1))
                    ws['A'+str(t_count+1)].style = st
                    ws['C'+str(t_count+1)].number_format = '0.00'
                    ws['D'+str(t_count+1)].number_format = '0.00'
                    ws['E'+str(t_count+1)].number_format = '0.00'
                    ws['F'+str(t_count+1)].number_format = '0.00'
                    ws['G'+str(t_count+1)].number_format = '0.00'
                    ws['H'+str(t_count+1)].number_format = '0.00'
                    ws['I'+str(t_count+1)].number_format = '0.00'
                    ws['J'+str(t_count+1)].number_format = '0.00'
                    ws['K'+str(t_count+1)].number_format = '0.00'
                    ws['L'+str(t_count+1)].number_format = '0.00'
                    ws['M'+str(t_count+1)].number_format = '0.00'
                    ws['N'+str(t_count+1)].number_format = '0.00'
                    ws['O'+str(t_count+1)].number_format = '0.00'
                    ws.row_dimensions[t_count+1].height = 22
                    t_count += 1

                    #Balance For Account Type
                    report_acc_obj = self.env['account.financial.report'].search([('parent_id','=',view_report.id),('type','=','account_type'),('sign','=','-1')])
                    for report_acc in report_acc_obj:
                        rep_account_ids = self.env['account.account'].search(
                            [('user_type_id', 'in', [x.id for x in report_acc.account_type_ids])])
                        acc_jan_bal, acc_feb_bal, acc_mar_bal, acc_apr_bal, acc_may_bal, acc_jun_bal, acc_july_bal = 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00
                        acc_aug_bal, acc_sep_bal, acc_oct_bal, acc_nov_bal, acc_dec_bal = 0.00, 0.00, 0.00, 0.00, 0.00
                        for acc in rep_account_ids:
                            year_from = report_ob.date_from[:4]
                            year_to = report_ob.date_to[:4]

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-01-01'),('date','<=',year_to+'-01-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_jan_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            if report_ob.date_to == year_to+'-02-29':
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-29'),('account_id','=',acc.id)]+domain)
                            else:
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-28'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_feb_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-03-01'),('date','<=',year_to+'-03-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_mar_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-04-01'),('date','<=',year_to+'-04-30'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_apr_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-05-01'),('date','<=',year_to+'-05-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_may_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-06-01'),('date','<=',year_to+'-06-30'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_jun_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-07-01'),('date','<=',year_to+'-07-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_july_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-08-01'),('date','<=',year_to+'-08-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_aug_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-09-01'),('date','<=',year_to+'-09-30'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_sep_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-10-01'),('date','<=',year_to+'-10-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_oct_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-11-01'),('date','<=',year_to+'-11-30'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_nov_bal += debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-12-01'),('date','<=',year_to+'-12-31'),('account_id','=',acc.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            acc_dec_bal += debit_bal - credit_bal

                        tot_bal = acc_jan_bal + acc_feb_bal + acc_mar_bal + acc_apr_bal + acc_may_bal + acc_jun_bal + acc_july_bal + acc_aug_bal + acc_sep_bal + acc_oct_bal + acc_nov_bal + acc_dec_bal

                        ws.append(['   ' + report_acc.name, '', acc_jan_bal, acc_feb_bal, acc_mar_bal, acc_apr_bal, acc_may_bal, acc_jun_bal, acc_july_bal, acc_aug_bal, acc_sep_bal, acc_oct_bal, acc_nov_bal, acc_dec_bal, tot_bal])
                        ft = Font(size=13, bold=True)
                        st = Style(font=ft)
                        ws.merge_cells("A" + str(t_count + 1) + ":B" + str(t_count + 1))
                        ws['A' + str(t_count + 1)].style = st
                        ws['C' + str(t_count + 1)].number_format = '0.00'
                        ws['D' + str(t_count + 1)].number_format = '0.00'
                        ws['E' + str(t_count + 1)].number_format = '0.00'
                        ws['F' + str(t_count + 1)].number_format = '0.00'
                        ws['G' + str(t_count + 1)].number_format = '0.00'
                        ws['H' + str(t_count + 1)].number_format = '0.00'
                        ws['I' + str(t_count + 1)].number_format = '0.00'
                        ws['J' + str(t_count + 1)].number_format = '0.00'
                        ws['K' + str(t_count + 1)].number_format = '0.00'
                        ws['L' + str(t_count + 1)].number_format = '0.00'
                        ws['M' + str(t_count + 1)].number_format = '0.00'
                        ws['N' + str(t_count + 1)].number_format = '0.00'
                        ws['O' + str(t_count + 1)].number_format = '0.00'
                        ws.row_dimensions[t_count + 1].height = 22
                        t_count += 1

                        #Month Wise balance For Accounts
                        account_ids = self.env['account.account'].search([('user_type_id','in', [x.id for x in report_acc.account_type_ids])])
                        for account in account_ids:
                            move_line_ids = self.env['account.move.line'].search([('account_id','=',account.id)]+domain)
                            account_bal = 0.00
                            for move in move_line_ids:
                                account_bal += move.debit - move.credit

                            year_from = report_ob.date_from[:4]
                            year_to = report_ob.date_to[:4]

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-01-01'),('date','<=',year_to+'-01-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            jan_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            if report_ob.date_to == year_to+'-02-29':
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-29'),('account_id','=',account.id)]+domain)
                            else:
                                move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-02-01'),('date','<=',year_to+'-02-28'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            feb_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-03-01'),('date','<=',year_to+'-03-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            mar_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-04-01'),('date','<=',year_to+'-04-30'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            apr_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-05-01'),('date','<=',year_to+'-05-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            may_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-06-01'),('date','<=',year_to+'-06-30'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            jun_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-07-01'),('date','<=',year_to+'-07-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            july_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-08-01'),('date','<=',year_to+'-08-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            aug_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-09-01'),('date','<=',year_to+'-09-30'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            sep_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-10-01'),('date','<=',year_to+'-10-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            oct_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-11-01'),('date','<=',year_to+'-11-30'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            nov_bal = debit_bal - credit_bal

                            debit_bal,credit_bal=0.00,0.00
                            move_line_ids = self.env['account.move.line'].search([('date','>=',year_from+'-12-01'),('date','<=',year_to+'-12-31'),('account_id','=',account.id)]+domain)
                            for line in move_line_ids:
                                debit_bal += line.debit
                                credit_bal += line.credit
                            dec_bal = debit_bal - credit_bal

                            ws.append(['       ' + account.name, '', jan_bal, feb_bal, mar_bal, apr_bal, may_bal, jun_bal,july_bal, aug_bal, sep_bal, oct_bal, nov_bal, dec_bal, account_bal])
                            ws.merge_cells("A" + str(t_count + 1) + ":B" + str(t_count + 1))
                            ws['C' + str(t_count + 1)].number_format = '0.00'
                            ws['D' + str(t_count + 1)].number_format = '0.00'
                            ws['E' + str(t_count + 1)].number_format = '0.00'
                            ws['F' + str(t_count + 1)].number_format = '0.00'
                            ws['G' + str(t_count + 1)].number_format = '0.00'
                            ws['H' + str(t_count + 1)].number_format = '0.00'
                            ws['I' + str(t_count + 1)].number_format = '0.00'
                            ws['J' + str(t_count + 1)].number_format = '0.00'
                            ws['K' + str(t_count + 1)].number_format = '0.00'
                            ws['L' + str(t_count + 1)].number_format = '0.00'
                            ws['M' + str(t_count + 1)].number_format = '0.00'
                            ws['N' + str(t_count + 1)].number_format = '0.00'
                            ws['O' + str(t_count + 1)].number_format = '0.00'
                            t_count += 1

                # Header1,2 Style
                ws.row_dimensions[1].height = 22
                ft = Font(size=14,bold=True,color='A901DB')
                st = Style(font=ft)
                ws['A1'].style = st
                ws['A1'].alignment = Alignment(horizontal="center")
                ws.merge_cells('A1:E1')
                ws.row_dimensions[2].height = 25
                ft = Font(size=18,bold=True,color='A901DB')
                st = Style(font=ft)
                ws['A2'].style = st
                ws['A2'].alignment = Alignment(horizontal="center")
                ws.merge_cells('A2:E2')

                #Header3 style
                ft = Font(size=12,bold=True)
                st = Style(font=ft)
                ws['A3'].style = st
                ws['A3'].alignment = Alignment(horizontal="center")
                ws['B3'].style = st
                ws['B3'].alignment = Alignment(horizontal="center")
                ws['C3'].style = st
                ws['C3'].alignment = Alignment(horizontal="center")
                ws.merge_cells('C3:D3')
                thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=3, column=1).border = thin_border
                ws.cell(row=3, column=2).border = thin_border
                ws.cell(row=3, column=3).border = thin_border
                ws.cell(row=3, column=4).border = thin_border

                # Header4,5 style
                ws['A4'].alignment = Alignment(horizontal="center")
                ws.merge_cells('A4:A5')
                ws['C4'].alignment = Alignment(horizontal="center")
                ws.merge_cells('C4:C5')
                ws['D4'].alignment = Alignment(horizontal="center")
                ws.merge_cells('D4:D5')
                thin_border = Border(right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=4, column=1).border = thin_border
                ws.cell(row=4, column=2).border = thin_border
                ws.cell(row=4, column=3).border = thin_border
                ws.cell(row=4, column=4).border = thin_border
                ws['A5'].alignment = Alignment(horizontal="center")
                ws['C5'].alignment = Alignment(horizontal="center")
                ws['D5'].alignment = Alignment(horizontal="center")

                thin_border = Border(bottom=Side(style='thin'),right=Side(style='thin'),left=Side(style='thin'))
                ws.cell(row=5, column=1).border = thin_border
                ws.cell(row=5, column=2).border = thin_border
                ws.cell(row=5, column=3).border = thin_border
                ws.cell(row=5, column=4).border = thin_border

                wb.save(excelreport_path)
                excel_file_obj = open(excelreport_path,'rb')
                bin_data= excel_file_obj.read()
                encoded_excel_data = base64.encodestring(bin_data)
                self.write({'file':encoded_excel_data, 'filename':'Statement_Of_Cash_Flow_Report.xlsx'})
                if excelreport_path:
                   try:
                       os.unlink(excelreport_path)
                   except (OSError, IOError):
                       _logger.error('Error when trying to remove file %s' % excelreport_path)

                return {'type': 'ir.actions.act_window',
                       'name':'Generate Excel Report',
                       'view_mode': 'form',
                       'view_type': 'form',
                       'res_id': self.id,
                       'res_model': 'cash.flow.statement',
                       'target': 'new',
                       'context': context,
                }

CashFlowStatement()